"""
XLSX Report Generator — Produces a professional Excel workbook with
Securitybricks branding for FedRAMP compliance scan results.

Includes Summary tab, Findings tab, and per-domain tabs with conditional
formatting and auto-adjusted column widths.
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
BRAND_NAVY = "0D4F4F"
BRAND_BLUE = "0E8585"
WHITE = "FFFFFF"
LIGHT_BG = "E8F5F5"
MET_GREEN = "28A745"
NOT_MET_RED = "DC3545"
MANUAL_AMBER = "FFC107"
ERROR_GRAY = "6C757D"
TEXT_DARK = "1A1A1A"

HEADER_FILL = PatternFill(start_color=BRAND_NAVY, end_color=BRAND_NAVY, fill_type="solid")
HEADER_FONT = Font(name="Aptos", bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name="Aptos", bold=True, color=BRAND_NAVY, size=14)
SUBTITLE_FONT = Font(name="Aptos", bold=True, color=BRAND_NAVY, size=11)
BODY_FONT = Font(name="Aptos", color=TEXT_DARK, size=10)
ALT_ROW_FILL = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D0D0D0"),
)

# Status fills
STATUS_FILLS = {
    "met": PatternFill(start_color=MET_GREEN, end_color=MET_GREEN, fill_type="solid"),
    "not_met": PatternFill(start_color=NOT_MET_RED, end_color=NOT_MET_RED, fill_type="solid"),
    "manual": PatternFill(start_color=MANUAL_AMBER, end_color=MANUAL_AMBER, fill_type="solid"),
    "error": PatternFill(start_color=ERROR_GRAY, end_color=ERROR_GRAY, fill_type="solid"),
}
STATUS_FONTS = {
    "met": Font(name="Aptos", bold=True, color=WHITE, size=10),
    "not_met": Font(name="Aptos", bold=True, color=WHITE, size=10),
    "manual": Font(name="Aptos", bold=True, color=TEXT_DARK, size=10),
    "error": Font(name="Aptos", bold=True, color=WHITE, size=10),
}

# Environment display names
ENV_NAMES = {
    "aws_commercial": "AWS Commercial",
    "aws_govcloud": "AWS GovCloud",
    "azure_commercial": "Azure Commercial",
    "azure_government": "Azure Government",
    "gcp_commercial": "GCP Commercial",
    "gcp_assured_workloads": "GCP Assured Workloads",
}


def generate_xlsx_report(scan, findings, client) -> bytes:
    """
    Generate a professional XLSX compliance report.

    Args:
        scan: Scan ORM object.
        findings: List of Finding ORM objects.
        client: Client ORM object.

    Returns:
        Bytes of the .xlsx file.
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Aggregate data
    status_counts = {"met": 0, "not_met": 0, "manual": 0, "error": 0}
    domain_data: dict[str, dict[str, Any]] = {}
    domain_findings: dict[str, list] = defaultdict(list)

    for f in findings:
        status_counts[f.status] = status_counts.get(f.status, 0) + 1

        if f.domain not in domain_data:
            domain_data[f.domain] = {
                "name": f.family,
                "met": 0,
                "not_met": 0,
                "manual": 0,
                "error": 0,
                "total": 0,
                "total_objectives": 0,
                "covered_objectives": 0,
            }
        domain_data[f.domain][f.status] = domain_data[f.domain].get(f.status, 0) + 1
        domain_data[f.domain]["total"] += 1
        domain_findings[f.domain].append(f)

        # Aggregate objective coverage
        cov = getattr(f, "objective_coverage", None) or {}
        if isinstance(cov, dict) and cov.get("total_objectives"):
            domain_data[f.domain]["total_objectives"] += cov["total_objectives"]
            domain_data[f.domain]["covered_objectives"] += cov["covered_objectives"]

    total = sum(status_counts.values())
    compliance_pct = round((status_counts["met"] / total * 100), 1) if total > 0 else 0.0

    # 1. Summary tab
    _create_summary_sheet(wb, scan, client, status_counts, compliance_pct, domain_data, total)

    # 2. All Findings tab
    _create_findings_sheet(wb, findings)

    # 3. Objective Coverage tab
    _create_coverage_sheet(wb, findings)

    # 4. Per-domain tabs
    for domain_code in sorted(domain_findings.keys()):
        d_findings = domain_findings[domain_code]
        d_name = domain_data[domain_code]["name"]
        _create_domain_sheet(wb, domain_code, d_name, d_findings)

    # Write to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _auto_width(ws, min_width: int = 10, max_width: int = 60):
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = adjusted


def _write_header_row(ws, row: int, headers: list[str]):
    """Write a formatted header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_data_row(ws, row: int, values: list, status_col: int = -1):
    """Write a data row with alternating fills and optional status formatting."""
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Alternating row shading
        if row % 2 == 0:
            cell.fill = ALT_ROW_FILL

        # Status column formatting
        if col == status_col and value in STATUS_FILLS:
            cell.fill = STATUS_FILLS[value]
            cell.font = STATUS_FONTS[value]
            cell.alignment = Alignment(horizontal="center", vertical="top")


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _create_summary_sheet(wb, scan, client, status_counts, compliance_pct, domain_data, total):
    ws = wb.create_sheet("Summary")

    # Title section
    row = 1
    ws.cell(row=row, column=1, value="FedRAMP Compliance Assessment Report").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    row = 2
    ws.cell(row=row, column=1, value="Securitybricks Cloud Compliance Scanner").font = Font(
        name="Aptos", italic=True, color=BRAND_BLUE, size=10
    )

    # Client info
    row = 4
    info_items = [
        ("Organization", client.name),
        ("Environment", ENV_NAMES.get(client.environment, client.environment)),
        ("FedRAMP Baseline", scan.fedramp_baseline),
        ("Scan ID", scan.id),
        ("Scan Date", scan.started_at.strftime("%Y-%m-%d %H:%M UTC") if scan.started_at else "N/A"),
        ("Completed", scan.completed_at.strftime("%Y-%m-%d %H:%M UTC") if scan.completed_at else "N/A"),
        ("Report Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for label, value in info_items:
        ws.cell(row=row, column=1, value=label).font = Font(name="Aptos", bold=True, color=BRAND_NAVY, size=10)
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        row += 1

    # Overall summary
    row += 1
    ws.cell(row=row, column=1, value="Overall Compliance Summary").font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    summary_headers = ["Total Checks", "Met", "Not Met", "Manual Review", "Error", "Compliance %"]
    _write_header_row(ws, row, summary_headers)
    row += 1

    summary_values = [
        total,
        status_counts["met"],
        status_counts["not_met"],
        status_counts["manual"],
        status_counts["error"],
        f"{compliance_pct}%",
    ]
    _write_data_row(ws, row, summary_values)

    # Apply status coloring to individual count cells
    ws.cell(row=row, column=2).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    ws.cell(row=row, column=3).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    ws.cell(row=row, column=4).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    ws.cell(row=row, column=5).fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")

    # Per-domain summary
    row += 2
    ws.cell(row=row, column=1, value="Per-Domain Breakdown").font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    domain_headers = ["Domain", "Family", "Met", "Not Met", "Manual", "Error", "Total", "Compliance %", "Obj. Covered", "Obj. Total", "Obj. Coverage %"]
    _write_header_row(ws, row, domain_headers)
    row += 1

    for code in sorted(domain_data.keys()):
        data = domain_data[code]
        d_total = data["total"]
        d_pct = round((data["met"] / d_total * 100), 1) if d_total > 0 else 0.0
        d_obj_total = data.get("total_objectives", 0)
        d_obj_covered = data.get("covered_objectives", 0)
        d_obj_pct = round((d_obj_covered / d_obj_total * 100), 1) if d_obj_total > 0 else 0.0
        values = [
            code,
            data["name"],
            data["met"],
            data["not_met"],
            data["manual"],
            data["error"],
            d_total,
            f"{d_pct}%",
            d_obj_covered,
            d_obj_total,
            f"{d_obj_pct}%",
        ]
        _write_data_row(ws, row, values)
        row += 1

    _auto_width(ws)
    ws.sheet_properties.tabColor = BRAND_NAVY


# ---------------------------------------------------------------------------
# All Findings sheet
# ---------------------------------------------------------------------------

def _create_findings_sheet(wb, findings):
    ws = wb.create_sheet("All Findings")

    headers = [
        "Control ID", "Family", "Domain", "Check ID", "Check Name",
        "Status", "Severity", "Obj. Covered", "Obj. Total", "Coverage %",
        "Evidence", "Remediation",
    ]
    _write_header_row(ws, 1, headers)

    for idx, f in enumerate(findings, start=2):
        cov = getattr(f, "objective_coverage", None) or {}
        cov_total = cov.get("total_objectives", 0) if isinstance(cov, dict) else 0
        cov_covered = cov.get("covered_objectives", 0) if isinstance(cov, dict) else 0
        cov_pct = cov.get("coverage_pct", 0.0) if isinstance(cov, dict) else 0.0
        values = [
            f.control_id,
            f.family,
            f.domain,
            f.check_id,
            f.check_name,
            f.status,
            f.severity.upper(),
            cov_covered,
            cov_total,
            f"{cov_pct}%",
            f.evidence or "",
            f.remediation or "",
        ]
        _write_data_row(ws, idx, values, status_col=6)

    _auto_width(ws)
    ws.sheet_properties.tabColor = BRAND_BLUE


# ---------------------------------------------------------------------------
# Objective Coverage sheet
# ---------------------------------------------------------------------------

OBJ_STATUS_FILLS = {
    "met": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    "not_met": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    "documentation_required": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    "not_tested": PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"),
    "error": PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"),
}


def _create_coverage_sheet(wb, findings):
    """Create a sheet showing NIST 800-53A objective-level coverage."""
    ws = wb.create_sheet("Objective Coverage")

    # Title
    ws.cell(row=1, column=1, value="NIST 800-53A Assessment Objective Coverage").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="Per-control breakdown of assessment objectives coverage").font = Font(
        name="Aptos", italic=True, color=BRAND_BLUE, size=10
    )

    headers = [
        "Control ID", "Control Name", "Domain", "Control Status",
        "Obj. ID", "Objective Text", "Obj. Status",
    ]
    header_row = 4
    _write_header_row(ws, header_row, headers)

    row = header_row + 1
    for f in findings:
        cov = getattr(f, "objective_coverage", None) or {}
        if not isinstance(cov, dict):
            continue
        details = cov.get("objective_details", [])
        if not details:
            continue

        for obj in details:
            values = [
                f.control_id,
                (f.check_name or "")[:60],
                f.domain,
                f.status,
                obj.get("id", ""),
                obj.get("text", ""),
                obj.get("status", "unknown"),
            ]
            _write_data_row(ws, row, values, status_col=4)

            # Color the objective status cell
            obj_status = obj.get("status", "")
            status_cell = ws.cell(row=row, column=7)
            if obj_status in OBJ_STATUS_FILLS:
                status_cell.fill = OBJ_STATUS_FILLS[obj_status]
                if obj_status in ("met",):
                    status_cell.font = Font(name="Aptos", bold=True, color="155724", size=10)
                elif obj_status in ("not_met",):
                    status_cell.font = Font(name="Aptos", bold=True, color="721C24", size=10)
                elif obj_status in ("documentation_required",):
                    status_cell.font = Font(name="Aptos", bold=True, color="856404", size=10)
                else:
                    status_cell.font = Font(name="Aptos", bold=True, color=ERROR_GRAY, size=10)
            status_cell.alignment = Alignment(horizontal="center", vertical="top")

            row += 1

    _auto_width(ws)
    ws.sheet_properties.tabColor = BRAND_BLUE


# ---------------------------------------------------------------------------
# Per-domain sheets
# ---------------------------------------------------------------------------

def _create_domain_sheet(wb, domain_code: str, domain_name: str, findings: list):
    """Create a sheet for a specific FedRAMP control family."""
    sheet_name = f"{domain_code} - {domain_name}"
    # Sheet names max 31 chars
    if len(sheet_name) > 31:
        sheet_name = f"{domain_code} - {domain_name[:31 - len(domain_code) - 3]}"

    ws = wb.create_sheet(sheet_name)

    # Domain header
    ws.cell(row=1, column=1, value=f"{domain_code}: {domain_name}").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # Domain stats
    met = sum(1 for f in findings if f.status == "met")
    not_met = sum(1 for f in findings if f.status == "not_met")
    manual = sum(1 for f in findings if f.status == "manual")
    error = sum(1 for f in findings if f.status == "error")
    d_total = len(findings)
    d_pct = round((met / d_total * 100), 1) if d_total > 0 else 0.0

    ws.cell(row=2, column=1, value="Total").font = Font(name="Aptos", bold=True, size=10)
    ws.cell(row=2, column=2, value=d_total).font = BODY_FONT
    ws.cell(row=2, column=3, value="Met").font = Font(name="Aptos", bold=True, color=MET_GREEN, size=10)
    ws.cell(row=2, column=4, value=met).font = BODY_FONT
    ws.cell(row=2, column=5, value="Not Met").font = Font(name="Aptos", bold=True, color=NOT_MET_RED, size=10)
    ws.cell(row=2, column=6, value=not_met).font = BODY_FONT
    ws.cell(row=2, column=7, value="Compliance").font = Font(name="Aptos", bold=True, size=10)
    ws.cell(row=2, column=8, value=f"{d_pct}%").font = Font(name="Aptos", bold=True, color=BRAND_NAVY, size=11)

    # Findings table
    headers = [
        "Control ID", "Check ID", "Check Name", "Status",
        "Severity", "Evidence", "Remediation",
    ]
    header_row = 4
    _write_header_row(ws, header_row, headers)

    for idx, f in enumerate(findings, start=header_row + 1):
        values = [
            f.control_id,
            f.check_id,
            f.check_name,
            f.status,
            f.severity.upper(),
            f.evidence or "",
            f.remediation or "",
        ]
        _write_data_row(ws, idx, values, status_col=4)

    _auto_width(ws)

    # Color the tab based on compliance
    if d_pct >= 80:
        ws.sheet_properties.tabColor = MET_GREEN
    elif d_pct >= 50:
        ws.sheet_properties.tabColor = MANUAL_AMBER
    else:
        ws.sheet_properties.tabColor = NOT_MET_RED
