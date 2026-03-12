#!/usr/bin/env python3
"""
Generate FedRAMP Scanner CLI Check Reference — XLSX and HTML reports.

Reads all config/checks/*.json files and produces:
  - reports/fedramp_cli_check_reference.xlsx  (3 tabs: AWS, Azure, GCP)
  - reports/fedramp_cli_check_reference.html  (3-tab HTML page)

Each row contains: Control Family, Control ID, Check ID, Check Name,
Service, CLI Command, What to Look For, Severity, Assessment Objectives,
Remediation.
"""
from __future__ import annotations

import html
import json
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
CHECKS_DIR = PROJECT_ROOT / "config" / "checks"
REPORTS_DIR = PROJECT_ROOT / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

XLSX_OUT = REPORTS_DIR / "fedramp_cli_check_reference.xlsx"
HTML_OUT = REPORTS_DIR / "fedramp_cli_check_reference.html"
CONTROLS_JSON = PROJECT_ROOT / "config" / "nist_800_53_controls.json"

# Max chars for objective text before we show only the code (e.g. "[a]")
OBJ_TEXT_MAX = 100

# ---------------------------------------------------------------------------
# CLI command mapping helpers
# ---------------------------------------------------------------------------

def _aws_cli(api_call: str) -> str:
    """Convert a boto3-style api_call to the corresponding AWS CLI command."""
    call = api_call.strip().rstrip("()")

    # Direct multi-service patterns first
    multi_patterns = {
        "ec2/guardduty/cloudtrail/kms": (
            "aws ec2 describe-vpcs  |  aws guardduty list-detectors  |  "
            "aws cloudtrail describe-trails  |  aws kms list-keys  |  "
            "aws ec2 describe-flow-logs"
        ),
    }
    if call in multi_patterns:
        return multi_patterns[call]

    # Handle combined calls ("+", "/", " + ")
    if "+" in call or ("/" in call and "." in call.split("/")[0]):
        parts = re.split(r'\s*[+/]\s*', call)
        return "  |  ".join(_aws_cli(p.strip()) for p in parts if p.strip())

    # Standard form: service.method or service.method()
    m = re.match(r'^(\w[\w-]*)\.(\w+)$', call)
    if m:
        svc, method = m.group(1), m.group(2)
        svc_cli = svc.replace("_", "-")
        # Convert camelCase/snake_case method to kebab-case
        method_cli = re.sub(r'([a-z0-9])([A-Z])', r'\1-\2', method)
        method_cli = method_cli.replace("_", "-").lower()
        return f"aws {svc_cli} {method_cli}"

    # service.sub_method.action  (e.g. elbv2.describe_load_balancer_attributes)
    parts = call.split(".")
    if len(parts) >= 2:
        svc = parts[0].replace("_", "-")
        method = "-".join(parts[1:])
        method = re.sub(r'([a-z0-9])([A-Z])', r'\1-\2', method)
        method = method.replace("_", "-").lower()
        return f"aws {svc} {method}"

    return f"aws {call}"


def _azure_cli(api_call: str) -> str:
    """Convert an Azure SDK-style api_call to az CLI equivalent."""
    call = api_call.strip()

    # Exact-match combined calls (checked before generic '+' splitting)
    combined_map = {
        "network_watchers.list_all + nsgs.list_all":
            "az network watcher list  |  az network nsg list",
        "NetworkManagementClient.network_watchers.list_all + resources.list":
            "az network watcher list  |  az resource list",
        "resource_client.providers.get/auth_client.role_assignments.list":
            "az provider show --namespace Microsoft.Security  |  az role assignment list",
        "KeyVaultManagementClient.vaults.list + get":
            "az keyvault list  |  az keyvault show --name $VAULT_NAME",
        "automation.automation_account.list + automation.runbook.list_by_automation_account":
            "az automation account list  |  az automation runbook list",
        "sql.servers.list + sql.replication_links.list_by_database":
            "az sql server list  |  az sql db replica list-links",
        "sql.servers.list + sql.backup_short_term_retention_policies.get":
            "az sql server list  |  az sql db str-policy show",
    }
    if call in combined_map:
        return combined_map[call]

    # Handle remaining combined calls generically
    if "+" in call:
        parts = re.split(r'\s*\+\s*', call)
        return "  |  ".join(_azure_cli(p.strip()) for p in parts if p.strip())

    # Graph API patterns
    if call.startswith("graph.") or call.startswith("graph/"):
        resource = call.replace("graph.", "").replace("graph/", "")
        # Remove trailing .list / .get action words
        resource = re.sub(r'\.(list|get)$', '', resource)
        # Convert to URL path
        resource = resource.replace("_", "")
        # camelCase to camelCase path (keep as-is for Graph URLs)
        resource = resource.replace(".", "/")
        return f"az rest --method GET --url https://graph.microsoft.com/v1.0/{resource}"

    # Multi-service shorthand
    multi = {
        "network/keyvault/monitor/security": (
            "az network nsg list  |  az keyvault list  |  "
            "az monitor activity-log alert list  |  "
            "az security pricing list  |  az network watcher list"
        ),
    }
    if call in multi:
        return multi[call]

    # SDK-style: Client.resource.method
    # e.g. authorization.role_definitions.list -> az role definition list
    sdk_map = {
        "AuthorizationManagementClient.role_assignments.list_for_scope": "az role assignment list",
        "authorization.role_assignments": "az role assignment list",
        "authorization.role_definitions": "az role definition list",
        "network.azure_firewalls": "az network firewall list",
        "network.network_security_groups": "az network nsg list",
        "network.virtual_networks": "az network vnet list",
        "network.virtual_network_gateways": "az network vnet-gateway list",
        "network.virtual_network_gateway_connections": "az network vpn-connection list",
        "network.bastion_hosts": "az network bastion list",
        "network.web_application_firewall_policies": "az network application-gateway waf-policy list",
        "network.application_gateways": "az network application-gateway list",
        "network.virtual_network_peerings": "az network vnet peering list",
        "compute.virtual_machines": "az vm list",
        "compute.virtual_machine_extensions": "az vm extension list",
        "compute.disks": "az disk list",
        "storage.storage_accounts": "az storage account list",
        "web.web_apps": "az webapp list",
        "web.certificates": "az webapp config ssl list",
        "sql.encryption_protectors": "az sql server tde-key show",
        "keyvault.vaults": "az keyvault list",
        "monitor.activity_log_alerts": "az monitor activity-log alert list",
        "monitor.diagnostic_settings": "az monitor diagnostic-settings list",
        "operationalinsights.workspaces": "az monitor log-analytics workspace list",
        "security.pricings": "az security pricing list",
        "security.security_contacts": "az security contact list",
        "security.secure_scores": "az security secure-score list",
        "security.assessments": "az security assessment list",
        "security.sub_assessments": "az security sub-assessment list",
        "security.jit_network_access_policies": "az security jit-policy list",
        "policy.policy_assignments": "az policy assignment list",
        "resources.management_locks": "az lock list",
        "advisor.recommendations": "az advisor recommendation list",
        "automation.automation_accounts": "az automation account list",
        "securityinsight.sentinel_onboarding_states": "az sentinel onboarding-state list",
        "resourcegraph.resources": "az graph query -q 'Resources'",
        "policy_assignments": "az policy assignment list",
        "diagnostic_settings": "az monitor diagnostic-settings list",
        "network_watchers": "az network watcher list",
        "storage_accounts": "az storage account list",
        "assessments": "az security assessment list",
        "sentinel_onboarding_states": "az sentinel onboarding-state list",
    }

    for prefix, cmd in sdk_map.items():
        if call.startswith(prefix):
            suffix = call[len(prefix):]
            if suffix and suffix.startswith("."):
                action = suffix.lstrip(".").replace("_", "-")
                action = re.sub(r'([a-z])([A-Z])', r'\1-\2', action).lower()
                return f"{cmd.rsplit(' ', 1)[0]} {action}"
            return cmd

    # Longer qualified names: MonitorManagementClient.*, etc.
    if "Client." in call or "ManagementClient." in call:
        simplified = re.sub(r'\w+Client\.', '', call)
        parts = simplified.split(".")
        resource = parts[0].replace("_", "-")
        resource = re.sub(r'([a-z])([A-Z])', r'\1-\2', resource).lower()
        action = parts[-1].replace("_", "-") if len(parts) > 1 else "list"
        action = re.sub(r'([a-z])([A-Z])', r'\1-\2', action).lower()
        return f"az {resource} {action}"

    # Sentinel
    if "sentinel" in call.lower():
        return "az sentinel onboarding-state list"

    # ResourceManagement
    if "ResourceManagement" in call or "resource_client" in call:
        return "az provider show --namespace Microsoft.Security"

    # Fallback
    clean = call.replace(".", " ").replace("_", "-")
    clean = re.sub(r'([a-z])([A-Z])', r'\1-\2', clean).lower()
    return f"az {clean}"


def _gcp_cli(api_call: str) -> str:
    """Convert a GCP REST-style api_call to gcloud CLI equivalent."""
    call = api_call.strip()

    # Handle combined / multi calls
    if "+" in call or ("/" in call and "." not in call.split("/")[0]):
        parts = re.split(r'\s*[+/]\s*', call)
        return "  |  ".join(_gcp_cli(p.strip()) for p in parts if p.strip())

    # Multi-service shorthand
    multi = {
        "compute/kms/logging/orgpolicy": (
            "gcloud compute firewall-rules list  |  gcloud kms keys list  |  "
            "gcloud logging sinks list  |  gcloud org-policies list"
        ),
    }
    if call in multi:
        return multi[call]

    gcp_map = {
        "cloudresourcemanager.projects.getIamPolicy": "gcloud projects get-iam-policy $PROJECT_ID",
        "iam.projects.serviceAccounts.list": "gcloud iam service-accounts list",
        "iam.projects.serviceAccounts.keys.list": "gcloud iam service-accounts keys list --iam-account=$SA_EMAIL",
        "iam.projects.roles.list": "gcloud iam roles list --project=$PROJECT_ID",
        "compute.instances.list": "gcloud compute instances list",
        "compute.instances.aggregatedList": "gcloud compute instances list",
        "compute.firewalls.list": "gcloud compute firewall-rules list",
        "compute.subnetworks.list": "gcloud compute networks subnets list",
        "compute.networks.listPeering": "gcloud compute networks peerings list",
        "compute.vpnTunnels.list": "gcloud compute vpn-tunnels list",
        "compute.vpnGateways.list": "gcloud compute vpn-gateways list",
        "compute.securityPolicies.list": "gcloud compute security-policies list",
        "compute.sslPolicies.list": "gcloud compute ssl-policies list",
        "compute.sslCertificates.list": "gcloud compute ssl-certificates list",
        "compute.backendServices.list": "gcloud compute backend-services list",
        "compute.packetMirrorings.list": "gcloud compute packet-mirrorings list",
        "compute.images.getIamPolicy": "gcloud compute images get-iam-policy $IMAGE_NAME",
        "compute.projects.get": "gcloud compute project-info describe",
        "logging.projects.logs.list": "gcloud logging logs list",
        "logging.projects.sinks.list": "gcloud logging sinks list",
        "logging.projects.locations.buckets.list": "gcloud logging buckets list",
        "logging.entries.list": "gcloud logging read 'logName:cloudaudit.googleapis.com' --limit=10",
        "monitoring.projects.alertPolicies.list": "gcloud alpha monitoring policies list",
        "storage.buckets.get": "gcloud storage buckets describe gs://$BUCKET_NAME",
        "storage.buckets.getIamPolicy": "gcloud storage buckets get-iam-policy gs://$BUCKET_NAME",
        "storage.buckets.list": "gcloud storage buckets list",
        "sqladmin.instances.list": "gcloud sql instances list",
        "container.projects.locations.clusters.list": "gcloud container clusters list",
        "containeranalysis.projects.occurrences.list": "gcloud artifacts docker images list --show-occurrences",
        "admin.directory.users.list": "gcloud identity groups memberships list (or Workspace Admin SDK)",
        "orgpolicy.projects.policies.get": "gcloud org-policies describe $CONSTRAINT --project=$PROJECT_ID",
        "orgpolicy.projects.policies.list": "gcloud org-policies list --project=$PROJECT_ID",
        "binaryauthorization.projects.getPolicy": "gcloud container binauthz policy export",
        "recommender.projects.locations.recommenders.recommendations.list": (
            "gcloud recommender recommendations list "
            "--recommender=google.iam.policy.Recommender --location=global"
        ),
        "securitycenter.securityHealthAnalyticsSettings": (
            "gcloud scc settings describe --organization=$ORG_ID"
        ),
        "securitycenter.organizations.notificationConfigs.list": (
            "gcloud scc notifications list --organization=$ORG_ID"
        ),
        "ids.projects.locations.endpoints.list": "gcloud ids endpoints list --location=$LOCATION",
        "websecurityscanner.projects.scanConfigs.list": "gcloud alpha web-security-scanner scan-configs list",
        "osconfig.projects.patchDeployments.list": "gcloud compute os-config patch-deployments list",
        "cloudasset.assets.list": "gcloud asset list --project=$PROJECT_ID",
        "cloudkms.projects.locations.keyRings.cryptoKeys.list": (
            "gcloud kms keys list --keyring=$KEYRING --location=$LOCATION"
        ),
        "cloudkms.projects.locations.keyRings.getIamPolicy": (
            "gcloud kms keyrings get-iam-policy $KEYRING --location=$LOCATION"
        ),
        "bigquery.datasets.list": "bq ls --project_id=$PROJECT_ID",
        "cloudresourcemanager.liens.list": "gcloud alpha resource-manager liens list",
    }
    if call in gcp_map:
        return gcp_map[call]

    # KMS key rotation via kms_v1
    if "kms_v1" in call or "KeyManagementService" in call:
        return "gcloud kms keys list --keyring=$KEYRING --location=$LOCATION --format='table(name,rotationPeriod,nextRotationTime)'"

    # BeyondCorp / session management
    if "BeyondCorp" in call or "session" in call.lower():
        return "gcloud alpha iap settings get --project=$PROJECT_ID"

    # orgpolicy shorthand
    if call.startswith("orgpolicy"):
        parts = call.split(".")
        return f"gcloud org-policies list --project=$PROJECT_ID"

    # Generic fallback: try to build from dot-separated path
    parts = call.split(".")
    if len(parts) >= 2:
        svc = parts[0]
        resource = parts[-2] if len(parts) > 2 else parts[0]
        action = parts[-1]
        resource = re.sub(r'([a-z])([A-Z])', r'\1-\2', resource).lower()
        action = re.sub(r'([a-z])([A-Z])', r'\1-\2', action).lower()
        return f"gcloud {resource} {action}"

    return f"gcloud {call}"


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

# Canonical family order
FAMILY_ORDER = [
    "AC", "AT", "AU", "CA", "CM", "CP", "IA", "IR",
    "MA", "MP", "PE", "PL", "PM", "PS", "PT", "RA",
    "SA", "SC", "SI", "SR",
]


def load_objective_texts() -> dict[str, dict[str, str]]:
    """Build {control_id: {code: text}} from nist_800_53_controls.json."""
    with open(CONTROLS_JSON) as f:
        data = json.load(f)

    result: dict[str, dict[str, str]] = {}
    for family_data in data.get("families", {}).values():
        for ctrl_id, ctrl in family_data.get("controls", {}).items():
            objs = ctrl.get("objectives", {})
            if objs:
                result[ctrl_id] = {code: o["text"] for code, o in objs.items()}
            for enh_key, enh in ctrl.get("enhancements", {}).items():
                # Convert "AC-2.9" → "AC-2(9)"
                enh_id = re.sub(r'\.(\d+)$', r'(\1)', enh_key)
                enh_objs = enh.get("objectives", {})
                if enh_objs:
                    result[enh_id] = {code: o["text"] for code, o in enh_objs.items()}
    return result


def _fmt_objective(code: str, text: str) -> str:
    """Format an objective: '[a] short text' or just '[a]' if text is too long."""
    if not text or len(text) > OBJ_TEXT_MAX:
        return code
    return f"{code} {text}"


def load_checks() -> dict[str, list[dict]]:
    """Return {'aws': [...], 'azure': [...], 'gcp': [...]}."""
    obj_texts = load_objective_texts()
    result: dict[str, list[dict]] = {"aws": [], "azure": [], "gcp": []}

    for json_path in sorted(CHECKS_DIR.glob("*.json")):
        with open(json_path) as f:
            data = json.load(f)
        domain = data["domain"]
        family_name = data["name"]

        for ctrl_id, ctrl_data in data.get("checks", {}).items():
            if ctrl_data.get("manual_only"):
                continue
            for csp in ("aws", "azure", "gcp"):
                for chk in ctrl_data.get(csp, []):
                    obj_codes = chk.get("supports_objectives", [])
                    ctrl_objs = obj_texts.get(ctrl_id, {})
                    obj_formatted = [
                        _fmt_objective(c, ctrl_objs.get(c, ""))
                        for c in obj_codes
                    ]
                    result[csp].append({
                        "family_code": domain,
                        "family_name": family_name,
                        "control_id": ctrl_id,
                        "check_id": chk["check_id"],
                        "check_name": chk["name"],
                        "service": chk.get("service", ""),
                        "objectives": obj_formatted,
                        "obj_codes": obj_codes,
                        "api_call": chk.get("api_call", ""),
                        "expected": chk.get("expected", ""),
                        "severity": chk.get("severity", ""),
                        "remediation": chk.get("remediation", ""),
                    })

    # Sort by family order then control_id
    def sort_key(row):
        fam = row["family_code"]
        idx = FAMILY_ORDER.index(fam) if fam in FAMILY_ORDER else 99
        # Extract numeric part for sorting controls: AC-2, AC-2(9), AC-3(8) etc.
        ctrl = row["control_id"]
        nums = re.findall(r'\d+', ctrl)
        num_tuple = tuple(int(n) for n in nums) if nums else (999,)
        return (idx, fam, num_tuple, row["check_id"])

    for csp in result:
        result[csp].sort(key=sort_key)

    return result


# ---------------------------------------------------------------------------
# CLI command conversion per CSP
# ---------------------------------------------------------------------------

CLI_CONVERTERS = {
    "aws": _aws_cli,
    "azure": _azure_cli,
    "gcp": _gcp_cli,
}


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

HEADERS = [
    "Control Family",
    "Control ID",
    "Check ID",
    "Check Name",
    "Service",
    "CLI Command",
    "What to Look For (Expected JSON Output)",
    "Severity",
    "Assessment Objectives",
    "Remediation",
]

# Colors per severity
SEV_FILLS = {
    "critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "high":     PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "medium":   PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "low":      PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
}

TAB_COLORS = {"aws": "FF9900", "azure": "0078D4", "gcp": "4285F4"}

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COL_WIDTHS = [22, 14, 20, 42, 18, 60, 52, 12, 34, 52]


def write_xlsx(all_checks: dict[str, list[dict]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for csp in ("aws", "azure", "gcp"):
        tab_name = f"{csp.upper()} ({len(all_checks[csp])} Checks)"
        ws = wb.create_sheet(title=tab_name)
        ws.sheet_properties.tabColor = TAB_COLORS[csp]
        converter = CLI_CONVERTERS[csp]
        checks = all_checks[csp]

        # Header row
        for col_idx, hdr in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER

        # Data rows
        for row_idx, chk in enumerate(checks, 2):
            cli_cmd = converter(chk["api_call"])
            values = [
                f'{chk["family_code"]} — {chk["family_name"]}',
                chk["control_id"],
                chk["check_id"],
                chk["check_name"],
                chk["service"],
                cli_cmd,
                chk["expected"],
                chk["severity"].capitalize(),
                "\n".join(chk["objectives"]),
                chk["remediation"],
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = WRAP
                cell.border = THIN_BORDER
                cell.font = Font(name="Calibri", size=10)
                # Severity color
                if col_idx == 8:
                    sev_fill = SEV_FILLS.get(chk["severity"])
                    if sev_fill:
                        cell.fill = sev_fill
                # Monospace for CLI column
                if col_idx == 6:
                    cell.font = Font(name="Consolas", size=9)

        # Column widths
        for col_idx, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Freeze header
        ws.freeze_panes = "A2"
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(checks) + 1}"

    wb.save(XLSX_OUT)
    print(f"  XLSX: {XLSX_OUT}")


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------

SEV_CSS = {
    "critical": "background:#ffc7ce;color:#9c0006;",
    "high": "background:#ffeb9c;color:#9c6500;",
    "medium": "background:#c6efce;color:#006100;",
    "low": "background:#d9e1f2;color:#003399;",
}

CSP_META = {
    "aws": ("AWS", "#FF9900"),
    "azure": ("Azure", "#0078D4"),
    "gcp": ("GCP", "#4285F4"),
}


def _esc(text: str) -> str:
    return html.escape(str(text))


def write_html(all_checks: dict[str, list[dict]]) -> None:
    # Build dynamic labels from actual counts
    csp_labels = {}
    for csp in ("aws", "azure", "gcp"):
        label, color = CSP_META[csp]
        count = f"{len(all_checks[csp])} Checks"
        csp_labels[csp] = (label, count, color)

    parts: list[str] = []

    from datetime import datetime
    now = datetime.now().strftime("%B %d, %Y")
    total_checks = sum(len(all_checks[csp]) for csp in ("aws", "azure", "gcp"))

    parts.append(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>FedRAMP CLI Check Reference — Securitybricks</title>
<style>
:root{{--sb-navy:#1B3A5C;--sb-navy-dark:#0F2440;--sb-blue:#2E75B6;--sb-accent:#c9a227;--sb-red:#E31B23;--sb-bg:#F5F7FA;--sb-text:#1A1A1A;}}
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Segoe UI',system-ui,-apple-system,sans-serif;background:var(--sb-bg);color:var(--sb-text);line-height:1.6;}}
.header{{
  background:linear-gradient(135deg, var(--sb-navy-dark) 0%, var(--sb-navy) 40%, var(--sb-blue) 100%);
  color:#fff;padding:36px 48px;position:relative;overflow:hidden;
}}
.header::after{{
  content:'';position:absolute;top:-50%;right:-10%;width:400px;height:400px;
  background:radial-gradient(circle, rgba(201,162,39,0.15) 0%, transparent 70%);border-radius:50%;
}}
.header h1{{font-size:26px;font-weight:700;margin-bottom:6px;position:relative;z-index:1;}}
.header .accent{{color:var(--sb-accent);font-weight:700;}}
.header .subtitle{{font-size:14px;opacity:0.85;position:relative;z-index:1;}}
.header .brand{{font-weight:700;font-size:13px;margin-top:10px;position:relative;z-index:1;}}
.header .brand span{{color:var(--sb-red);}}
.header .generated{{font-size:12px;opacity:0.7;margin-top:4px;position:relative;z-index:1;}}
.tabs{{display:flex;gap:0;background:#f8f9fa;padding:0 32px;border-bottom:2px solid #dee2e6;position:sticky;top:0;z-index:100;}}
.tab{{padding:12px 28px;font-weight:600;font-size:14px;cursor:pointer;border:none;background:transparent;
  border-bottom:3px solid transparent;transition:all .15s;color:#666;}}
.tab:hover{{background:#e9ecef;}}
.tab.active{{border-bottom-color:var(--active-color);color:var(--active-color);background:#fff;}}
.tab .badge{{font-size:10px;background:#ddd;color:#555;border-radius:10px;padding:2px 7px;margin-left:6px;}}
.tab.active .badge{{background:var(--active-color);color:#fff;}}
.panel{{display:none;padding:16px 32px 32px;}}
.panel.active{{display:block;}}
.summary{{display:flex;gap:16px;margin-bottom:16px;flex-wrap:wrap;}}
.summary .stat{{
  background:#fff;border-radius:10px;padding:16px 20px;min-width:160px;text-align:center;
  box-shadow:0 2px 8px rgba(0,0,0,0.06);border-top:4px solid #ddd;transition:transform 0.2s;
}}
.summary .stat:hover{{transform:translateY(-2px);box-shadow:0 4px 16px rgba(0,0,0,0.1);}}
.summary .stat .num{{font-size:32px;font-weight:800;line-height:1.1;}}
.summary .stat .label{{font-size:12px;color:#666;text-transform:uppercase;letter-spacing:.5px;margin-top:4px;}}
.search-row{{margin-bottom:12px;}}
.search-row input{{padding:10px 14px;border:1px solid #ddd;border-radius:6px;width:320px;font-size:13px;transition:border-color 0.2s;}}
.search-row input:focus{{outline:none;border-color:var(--sb-blue);box-shadow:0 0 0 3px rgba(46,117,182,0.15);}}
table{{width:100%;border-collapse:collapse;font-size:12px;background:#fff;border:1px solid #dee2e6;border-radius:8px;overflow:hidden;}}
thead{{background:var(--sb-navy);}}
th{{color:#fff;padding:10px 8px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.4px;position:sticky;top:0;z-index:1;}}
td{{padding:8px;border-bottom:1px solid #eee;vertical-align:top;}}
tr:hover td{{background:#f0f4ff;}}
.cli{{font-family:'Consolas','Courier New',monospace;font-size:11px;background:#f4f4f4;padding:2px 4px;border-radius:3px;word-break:break-all;}}
.sev{{padding:2px 10px;border-radius:12px;font-size:10px;font-weight:700;text-transform:uppercase;display:inline-block;}}
.sev-critical{{background:#dc3545;color:#fff;}}
.sev-high{{background:#fd7e14;color:#fff;}}
.sev-medium{{background:#ffc107;color:#333;}}
.sev-low{{background:#28a745;color:#fff;}}
.family-group{{background:#edf2f7;font-weight:600;font-size:12px;}}
.footer{{text-align:center;padding:24px;font-size:11px;color:#888;border-top:1px solid #eee;margin:0 32px 0;}}
@media print{{.tabs,.search-row{{display:none;}}.panel{{display:block!important;page-break-before:always;}}}}
</style>
</head>
<body>
<div class="header">
  <h1><span class="accent">FedRAMP</span> Cloud Compliance Scanner &mdash; CLI Check Reference</h1>
  <div class="subtitle">NIST 800-53 Rev 5 Automated Controls &bull; AWS, Azure, GCP</div>
  <div class="brand">SECURITY<span>BRICKS</span> &mdash; Powered by Aprio (3PAO)</div>
  <div class="generated">Generated: {now} &bull; {total_checks} Total Checks</div>
</div>

<div class="tabs">
""")

    # Tab buttons
    for i, csp in enumerate(("aws", "azure", "gcp")):
        label, count, color = csp_labels[csp]
        active = " active" if i == 0 else ""
        parts.append(
            f'  <button class="tab{active}" style="--active-color:{color}" '
            f'onclick="switchTab(\'{csp}\')" data-csp="{csp}">'
            f'{label}<span class="badge">{count}</span></button>\n'
        )

    parts.append("</div>\n")

    # Panels
    for i, csp in enumerate(("aws", "azure", "gcp")):
        label, count, color = csp_labels[csp]
        converter = CLI_CONVERTERS[csp]
        checks = all_checks[csp]
        active = " active" if i == 0 else ""

        # Count families, unique objectives & severities
        families = set()
        objectives_set = set()
        sev_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for chk in checks:
            families.add(chk["family_code"])
            for obj_code in chk.get("obj_codes", []):
                objectives_set.add((chk["control_id"], obj_code))
            sev_counts[chk["severity"]] = sev_counts.get(chk["severity"], 0) + 1

        parts.append(f'<div class="panel{active}" id="panel-{csp}">\n')
        parts.append('<div class="summary">\n')
        parts.append(f'  <div class="stat" style="border-top-color:var(--sb-navy);"><div class="num" style="color:var(--sb-navy);">{len(checks)}</div><div class="label">Total Checks</div></div>\n')
        parts.append(f'  <div class="stat" style="border-top-color:var(--sb-blue);"><div class="num" style="color:var(--sb-blue);">{len(families)}</div><div class="label">Control Families</div></div>\n')
        parts.append(f'  <div class="stat" style="border-top-color:var(--sb-accent);"><div class="num" style="color:var(--sb-accent);">{len(objectives_set)}</div><div class="label">Control Objectives</div></div>\n')
        parts.append(f'  <div class="stat" style="border-top-color:#dc3545;"><div class="num" style="color:#dc3545;">{sev_counts.get("critical",0)}</div><div class="label">Critical</div></div>\n')
        parts.append(f'  <div class="stat" style="border-top-color:#fd7e14;"><div class="num" style="color:#fd7e14;">{sev_counts.get("high",0)}</div><div class="label">High</div></div>\n')
        parts.append(f'  <div class="stat" style="border-top-color:#ffc107;"><div class="num" style="color:#b8860b;">{sev_counts.get("medium",0)}</div><div class="label">Medium</div></div>\n')
        parts.append(f'  <div class="stat" style="border-top-color:#28a745;"><div class="num" style="color:#28a745;">{sev_counts.get("low",0)}</div><div class="label">Low</div></div>\n')
        parts.append('</div>\n')

        parts.append(f'<div class="search-row"><input type="text" placeholder="Filter checks…" oninput="filterTable(\'{csp}\', this.value)"></div>\n')

        parts.append(f'<table id="table-{csp}">\n<thead><tr>\n')
        for hdr in HEADERS:
            parts.append(f"  <th>{_esc(hdr)}</th>\n")
        parts.append("</tr></thead>\n<tbody>\n")

        prev_family = ""
        for chk in checks:
            family_label = f'{chk["family_code"]} — {chk["family_name"]}'
            if family_label != prev_family:
                parts.append(f'<tr class="family-group"><td colspan="10">{_esc(family_label)}</td></tr>\n')
                prev_family = family_label

            cli_cmd = converter(chk["api_call"])
            sev = chk["severity"]
            sev_class = f"sev-{sev}" if sev in SEV_CSS else ""

            obj_html = "<br>".join(_esc(o) for o in chk["objectives"])

            parts.append("<tr>\n")
            parts.append(f'  <td>{_esc(family_label)}</td>\n')
            parts.append(f'  <td><strong>{_esc(chk["control_id"])}</strong></td>\n')
            parts.append(f'  <td>{_esc(chk["check_id"])}</td>\n')
            parts.append(f'  <td>{_esc(chk["check_name"])}</td>\n')
            parts.append(f'  <td>{_esc(chk["service"])}</td>\n')
            parts.append(f'  <td><span class="cli">{_esc(cli_cmd)}</span></td>\n')
            parts.append(f'  <td>{_esc(chk["expected"])}</td>\n')
            parts.append(f'  <td><span class="sev {sev_class}">{_esc(sev.capitalize())}</span></td>\n')
            parts.append(f'  <td>{obj_html}</td>\n')
            parts.append(f'  <td>{_esc(chk["remediation"])}</td>\n')
            parts.append("</tr>\n")

        parts.append("</tbody></table>\n</div>\n")

    # Footer (f-string for dynamic values)
    parts.append(f"""
<div class="footer">
  FedRAMP Cloud Compliance Scanner &mdash; Securitybricks (Powered by Aprio) &mdash; For authorized assessment use only<br>
  {total_checks} checks across AWS, Azure, GCP &bull; Generated: {now}
</div>
""")

    # JS (regular string — curly braces must not be interpreted)
    parts.append("""
<script>
function switchTab(csp) {
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
  document.querySelector(`.tab[data-csp="${csp}"]`).classList.add('active');
  document.getElementById(`panel-${csp}`).classList.add('active');
}

function filterTable(csp, query) {
  const q = query.toLowerCase();
  const rows = document.querySelectorAll(`#table-${csp} tbody tr`);
  rows.forEach(row => {
    if (row.classList.contains('family-group')) {
      row.style.display = '';
      return;
    }
    const text = row.textContent.toLowerCase();
    row.style.display = text.includes(q) ? '' : 'none';
  });
  // Hide empty family headers
  let lastFamily = null;
  const allRows = Array.from(rows);
  for (let i = allRows.length - 1; i >= 0; i--) {
    const r = allRows[i];
    if (r.classList.contains('family-group')) {
      if (!lastFamily || lastFamily.style.display === 'none') {
        r.style.display = 'none';
      }
      lastFamily = null;
    } else {
      if (r.style.display !== 'none') lastFamily = r;
    }
  }
}
</script>
</body>
</html>
""")

    HTML_OUT.write_text("".join(parts), encoding="utf-8")
    print(f"  HTML: {HTML_OUT}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading checks from config/checks/*.json …")
    all_checks = load_checks()

    for csp in ("aws", "azure", "gcp"):
        print(f"  {csp.upper()}: {len(all_checks[csp])} checks")

    print("\nGenerating reports …")
    write_xlsx(all_checks)
    write_html(all_checks)
    print("\nDone.")


if __name__ == "__main__":
    main()
